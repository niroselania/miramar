from __future__ import annotations

import cgi
import io
import json
import os
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

try:
    import openpyxl
except Exception as exc:  # pragma: no cover - startup guard
    openpyxl = None
    OPENPYXL_ERROR = exc
else:
    OPENPYXL_ERROR = None


APP_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.environ.get("MIRAMAR_DB", APP_DIR / "miramar.db"))
EXPORT_DIR = Path(os.environ.get("MIRAMAR_EXPORT_DIR", APP_DIR / "exports"))
DEFAULT_EXCEL = Path(r"C:\Users\ARIEL-ROSETI\Downloads\Miramar\Servicio Miramar 2026.xlsx")

MONTHS = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]

CATEGORIES = ["EDEA", "GAS", "EXPENSAS", "MGA COCHERA", "MGA DPTO"]


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    EXPORT_DIR.mkdir(exist_ok=True)
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    with db() as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                year INTEGER NOT NULL,
                month INTEGER NOT NULL,
                month_name TEXT NOT NULL,
                category TEXT NOT NULL,
                amount REAL NOT NULL DEFAULT 0,
                notes TEXT NOT NULL DEFAULT '',
                updated_at TEXT NOT NULL,
                UNIQUE(year, month, category)
            );

            CREATE TABLE IF NOT EXISTS property_notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                label TEXT NOT NULL,
                value TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );
            """
        )


def now() -> str:
    return datetime.now().isoformat(timespec="seconds")


def money(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def infer_year(sheet_name: str) -> int:
    for token in sheet_name.replace("-", " ").split():
        if token.isdigit() and len(token) == 4:
            return int(token)
    return datetime.now().year


def import_workbook(data: bytes, source_name: str = "Excel cargado") -> dict[str, Any]:
    if openpyxl is None:
        raise RuntimeError(f"openpyxl no esta disponible: {OPENPYXL_ERROR}")

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    imported = 0
    notes = 0

    with db() as conn:
        for ws in wb.worksheets:
            headers = [cell.value for cell in ws[3]] if ws.max_row >= 3 else []
            if "Periodo" in headers and "EDEA" in headers:
                year = infer_year(ws.title)
                header_map = {str(v).strip(): idx + 1 for idx, v in enumerate(headers) if v}
                for row in range(4, ws.max_row + 1):
                    month_name = ws.cell(row, header_map["Periodo"]).value
                    if month_name not in MONTHS:
                        continue
                    month = MONTHS.index(month_name) + 1
                    for category in CATEGORIES:
                        col = header_map.get(category)
                        if not col:
                            continue
                        value = money(ws.cell(row, col).value)
                        conn.execute(
                            """
                            INSERT INTO expenses (year, month, month_name, category, amount, notes, updated_at)
                            VALUES (?, ?, ?, ?, ?, '', ?)
                            ON CONFLICT(year, month, category)
                            DO UPDATE SET amount = excluded.amount, updated_at = excluded.updated_at
                            """,
                            (year, month, month_name, category, value, now()),
                        )
                        imported += 1
            elif ws.title.lower().startswith("inmobiliario"):
                conn.execute("DELETE FROM property_notes")
                for row in ws.iter_rows(values_only=True):
                    values = [str(v).strip() for v in row if v not in (None, "")]
                    if not values:
                        continue
                    conn.execute(
                        "INSERT INTO property_notes (label, value, updated_at) VALUES (?, ?, ?)",
                        (values[0], " | ".join(values[1:]) if len(values) > 1 else "", now()),
                    )
                    notes += 1

    return {"source": source_name, "expenses": imported, "notes": notes}


def seed_default_if_empty() -> None:
    with db() as conn:
        count = conn.execute("SELECT COUNT(*) FROM expenses").fetchone()[0]
    if count == 0 and DEFAULT_EXCEL.exists():
        import_workbook(DEFAULT_EXCEL.read_bytes(), DEFAULT_EXCEL.name)


def rows_to_dicts(rows: list[sqlite3.Row]) -> list[dict[str, Any]]:
    return [dict(row) for row in rows]


def get_dashboard() -> dict[str, Any]:
    with db() as conn:
        expense_rows = rows_to_dicts(
            conn.execute(
                """
                SELECT * FROM expenses
                ORDER BY year DESC, month ASC,
                CASE category
                    WHEN 'EDEA' THEN 1
                    WHEN 'GAS' THEN 2
                    WHEN 'EXPENSAS' THEN 3
                    WHEN 'MGA COCHERA' THEN 4
                    WHEN 'MGA DPTO' THEN 5
                    ELSE 99
                END
                """
            ).fetchall()
        )
        notes = rows_to_dicts(conn.execute("SELECT * FROM property_notes ORDER BY id").fetchall())

    by_month: dict[str, dict[str, Any]] = {}
    by_category = {category: 0.0 for category in CATEGORIES}
    total = 0.0

    for row in expense_rows:
        key = f"{row['year']}-{row['month']:02d}"
        bucket = by_month.setdefault(
            key,
            {
                "year": row["year"],
                "month": row["month"],
                "month_name": row["month_name"],
                "categories": {category: 0.0 for category in CATEGORIES},
                "total": 0.0,
                "half": 0.0,
            },
        )
        amount = float(row["amount"] or 0)
        bucket["categories"][row["category"]] = amount
        bucket["total"] += amount
        bucket["half"] = bucket["total"] / 2
        by_category[row["category"]] = by_category.get(row["category"], 0) + amount
        total += amount

    months = sorted(by_month.values(), key=lambda item: (item["year"], item["month"]))
    years = sorted({row["year"] for row in expense_rows}, reverse=True)
    return {
        "expenses": expense_rows,
        "months": months,
        "categories": CATEGORIES,
        "byCategory": by_category,
        "total": total,
        "halfTotal": total / 2,
        "years": years,
        "notes": notes,
    }


def upsert_expense(payload: dict[str, Any]) -> dict[str, Any]:
    year = int(payload.get("year") or datetime.now().year)
    month = int(payload.get("month") or 1)
    category = str(payload.get("category") or "").strip().upper()
    if category not in CATEGORIES:
        raise ValueError("Categoria invalida")
    if not 1 <= month <= 12:
        raise ValueError("Mes invalido")
    amount = money(payload.get("amount"))
    notes = str(payload.get("notes") or "").strip()
    with db() as conn:
        conn.execute(
            """
            INSERT INTO expenses (year, month, month_name, category, amount, notes, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(year, month, category)
            DO UPDATE SET amount = excluded.amount, notes = excluded.notes, updated_at = excluded.updated_at
            """,
            (year, month, MONTHS[month - 1], category, amount, notes, now()),
        )
    return {"ok": True}


def delete_expense(expense_id: int) -> dict[str, Any]:
    with db() as conn:
        conn.execute("DELETE FROM expenses WHERE id = ?", (expense_id,))
    return {"ok": True}


def export_xlsx() -> Path:
    if openpyxl is None:
        raise RuntimeError(f"openpyxl no esta disponible: {OPENPYXL_ERROR}")

    data = get_dashboard()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Miramar"
    headers = ["Periodo", *CATEGORIES, "Total", "/2"]
    ws.append(["Servicios Dpto"])
    ws.append([])
    ws.append(headers)

    for item in data["months"]:
        ws.append(
            [
                item["month_name"],
                *[item["categories"].get(category, 0) for category in CATEGORIES],
                item["total"],
                item["half"],
            ]
        )

    total_row = ws.max_row + 2
    ws.cell(total_row, 1, "Total")
    for col in range(2, len(headers) + 1):
        letter = openpyxl.utils.get_column_letter(col)
        ws.cell(total_row, col, f"=SUM({letter}4:{letter}{ws.max_row - 2})")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=len(headers)):
        for cell in row:
            cell.number_format = '"$"#,##0'

    notes_ws = wb.create_sheet("inmobiliario")
    notes_ws.append(["Dato", "Detalle"])
    for note in data["notes"]:
        notes_ws.append([note["label"], note["value"]])
    notes_ws.column_dimensions["A"].width = 38
    notes_ws.column_dimensions["B"].width = 60

    out = EXPORT_DIR / f"miramar_gastos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out)
    return out


INDEX_HTML = r"""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Miramar - Gastos</title>
  <style>
    :root {
      color-scheme: light;
      --ink: #14213d;
      --muted: #617089;
      --line: #d8dee9;
      --panel: #ffffff;
      --wash: #f4f7fb;
      --accent: #0f766e;
      --accent-2: #b45309;
      --danger: #b42318;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: Inter, "Segoe UI", Arial, sans-serif;
      background: var(--wash);
      color: var(--ink);
    }
    header {
      background: #ffffff;
      border-bottom: 1px solid var(--line);
      padding: 18px clamp(16px, 4vw, 42px);
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      position: sticky;
      top: 0;
      z-index: 3;
    }
    h1 { margin: 0; font-size: clamp(22px, 3vw, 34px); letter-spacing: 0; }
    main { padding: 22px clamp(16px, 4vw, 42px) 46px; }
    .actions { display: flex; flex-wrap: wrap; gap: 10px; align-items: center; }
    button, .button, input, select {
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fff;
      color: var(--ink);
      min-height: 38px;
      padding: 8px 10px;
      font: inherit;
    }
    button, .button {
      cursor: pointer;
      text-decoration: none;
      display: inline-flex;
      align-items: center;
      gap: 7px;
      font-weight: 650;
    }
    button.primary, .button.primary { background: var(--accent); color: white; border-color: var(--accent); }
    button.warn { background: #fff8ed; border-color: #f3c278; color: var(--accent-2); }
    button.danger { background: #fff5f5; border-color: #f6b5ad; color: var(--danger); }
    .grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(160px, 1fr));
      gap: 12px;
      margin: 20px 0;
    }
    .metric {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
    }
    .metric span { display: block; color: var(--muted); font-size: 13px; }
    .metric strong { display: block; margin-top: 6px; font-size: 24px; }
    section {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      margin-top: 16px;
      overflow: hidden;
    }
    section h2 {
      margin: 0;
      padding: 14px 16px;
      border-bottom: 1px solid var(--line);
      font-size: 18px;
    }
    .section-body { padding: 16px; }
    .table-wrap { overflow-x: auto; }
    table { width: 100%; border-collapse: collapse; min-width: 860px; }
    th, td { padding: 10px 9px; border-bottom: 1px solid var(--line); text-align: right; white-space: nowrap; }
    th:first-child, td:first-child { text-align: left; }
    th { background: #f8fafc; color: #46556f; font-size: 13px; }
    tr:hover td { background: #fbfdff; }
    .form-grid {
      display: grid;
      grid-template-columns: 110px 145px minmax(150px, 1fr) 145px minmax(160px, 1fr) auto;
      gap: 10px;
      align-items: end;
    }
    label { display: grid; gap: 5px; color: var(--muted); font-size: 12px; font-weight: 650; }
    label input, label select { width: 100%; color: var(--ink); font-size: 14px; font-weight: 400; }
    .status { color: var(--muted); min-height: 22px; margin-top: 10px; }
    .notes { display: grid; gap: 8px; color: var(--muted); }
    .note-row { display: grid; grid-template-columns: minmax(160px, 260px) 1fr; gap: 12px; padding: 8px 0; border-bottom: 1px solid var(--line); }
    .note-row strong, .note-row span { color: var(--ink); overflow-wrap: anywhere; }
    .empty { color: var(--muted); padding: 18px; }
    @media (max-width: 900px) {
      header { align-items: flex-start; flex-direction: column; }
      .grid { grid-template-columns: repeat(2, minmax(140px, 1fr)); }
      .form-grid { grid-template-columns: 1fr 1fr; }
      .form-grid button { grid-column: span 2; justify-content: center; }
    }
    @media (max-width: 560px) {
      .grid { grid-template-columns: 1fr; }
      .form-grid { grid-template-columns: 1fr; }
      .form-grid button { grid-column: auto; }
      .note-row { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
  <header>
    <div>
      <h1>Gastos Miramar</h1>
      <div class="status" id="subtitle">Base local SQLite visible desde el navegador</div>
    </div>
    <div class="actions">
      <form id="uploadForm" enctype="multipart/form-data">
        <input id="file" name="file" type="file" accept=".xlsx,.xlsm">
        <button class="primary" type="submit">Cargar Excel</button>
      </form>
      <button id="exportBtn" type="button">Exportar Excel</button>
    </div>
  </header>
  <main>
    <div class="grid" id="metrics"></div>

    <section>
      <h2>Cargar o actualizar gasto</h2>
      <div class="section-body">
        <form id="expenseForm" class="form-grid">
          <label>Año<input name="year" type="number" min="2020" max="2100" value="2026"></label>
          <label>Mes<select name="month" id="monthSelect"></select></label>
          <label>Concepto<select name="category" id="categorySelect"></select></label>
          <label>Importe<input name="amount" type="number" min="0" step="0.01" placeholder="0"></label>
          <label>Nota<input name="notes" type="text" placeholder="Opcional"></label>
          <button class="primary" type="submit">Guardar</button>
        </form>
        <div class="status" id="formStatus"></div>
      </div>
    </section>

    <section>
      <h2>Resumen mensual</h2>
      <div class="table-wrap">
        <table>
          <thead id="monthHead"></thead>
          <tbody id="monthRows"></tbody>
        </table>
      </div>
    </section>

    <section>
      <h2>Registros de la base</h2>
      <div class="table-wrap">
        <table>
          <thead><tr><th>Año</th><th>Mes</th><th>Concepto</th><th>Importe</th><th>Nota</th><th>Actualizado</th><th></th></tr></thead>
          <tbody id="expenseRows"></tbody>
        </table>
      </div>
    </section>

    <section>
      <h2>Datos inmobiliarios</h2>
      <div class="section-body notes" id="notes"></div>
    </section>
  </main>
  <script>
    const months = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"];
    let categories = [];
    const money = new Intl.NumberFormat("es-AR", { style: "currency", currency: "ARS", maximumFractionDigits: 0 });
    const qs = (sel) => document.querySelector(sel);

    function setStatus(text) { qs("#formStatus").textContent = text || ""; }

    async function api(path, options) {
      const res = await fetch(path, options);
      if (!res.ok) {
        const body = await res.json().catch(() => ({}));
        throw new Error(body.error || "No se pudo completar la operacion");
      }
      return res.json();
    }

    function fillSelects(data) {
      categories = data.categories;
      qs("#monthSelect").innerHTML = months.map((m, i) => `<option value="${i + 1}">${m}</option>`).join("");
      qs("#categorySelect").innerHTML = categories.map(c => `<option>${c}</option>`).join("");
      if (data.years[0]) qs('input[name="year"]').value = data.years[0];
    }

    function renderMetrics(data) {
      const topCategory = Object.entries(data.byCategory).sort((a, b) => b[1] - a[1])[0] || ["-", 0];
      qs("#metrics").innerHTML = `
        <div class="metric"><span>Total cargado</span><strong>${money.format(data.total)}</strong></div>
        <div class="metric"><span>Mitad total</span><strong>${money.format(data.halfTotal)}</strong></div>
        <div class="metric"><span>Meses con datos</span><strong>${data.months.length}</strong></div>
        <div class="metric"><span>Mayor concepto</span><strong>${topCategory[0]} ${money.format(topCategory[1])}</strong></div>
      `;
    }

    function renderMonthTable(data) {
      qs("#monthHead").innerHTML = `<tr><th>Periodo</th>${data.categories.map(c => `<th>${c}</th>`).join("")}<th>Total</th><th>/2</th></tr>`;
      qs("#monthRows").innerHTML = data.months.map(row => `
        <tr>
          <td>${row.month_name} ${row.year}</td>
          ${data.categories.map(c => `<td>${money.format(row.categories[c] || 0)}</td>`).join("")}
          <td><strong>${money.format(row.total)}</strong></td>
          <td>${money.format(row.half)}</td>
        </tr>
      `).join("") || `<tr><td colspan="${data.categories.length + 3}" class="empty">Todavia no hay gastos cargados.</td></tr>`;
    }

    function renderExpenseRows(data) {
      qs("#expenseRows").innerHTML = data.expenses.map(row => `
        <tr>
          <td>${row.year}</td>
          <td>${row.month_name}</td>
          <td>${row.category}</td>
          <td>${money.format(row.amount)}</td>
          <td>${row.notes || ""}</td>
          <td>${row.updated_at}</td>
          <td><button class="danger" data-delete="${row.id}" type="button">Borrar</button></td>
        </tr>
      `).join("") || `<tr><td colspan="7" class="empty">La base esta vacia.</td></tr>`;
    }

    function renderNotes(data) {
      qs("#notes").innerHTML = data.notes.map(note => `
        <div class="note-row"><strong>${note.label}</strong><span>${note.value || ""}</span></div>
      `).join("") || `<div class="empty">No hay datos inmobiliarios cargados.</div>`;
    }

    async function load() {
      const data = await api("/api/data");
      fillSelects(data);
      renderMetrics(data);
      renderMonthTable(data);
      renderExpenseRows(data);
      renderNotes(data);
    }

    qs("#expenseForm").addEventListener("submit", async (event) => {
      event.preventDefault();
      const form = new FormData(event.currentTarget);
      await api("/api/expense", {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(Object.fromEntries(form.entries())),
      });
      event.currentTarget.reset();
      qs('input[name="year"]').value = new Date().getFullYear();
      setStatus("Gasto guardado.");
      await load();
    });

    qs("#expenseRows").addEventListener("click", async (event) => {
      const button = event.target.closest("[data-delete]");
      if (!button) return;
      await api(`/api/expense?id=${button.dataset.delete}`, { method: "DELETE" });
      setStatus("Registro borrado.");
      await load();
    });

    qs("#uploadForm").addEventListener("submit", async (event) => {
      event.preventDefault();
      const form = new FormData(event.currentTarget);
      setStatus("Importando Excel...");
      const res = await fetch("/api/import", { method: "POST", body: form });
      const body = await res.json();
      if (!res.ok) throw new Error(body.error || "No se pudo importar");
      setStatus(`Excel importado: ${body.expenses} importes y ${body.notes} notas.`);
      await load();
    });

    qs("#exportBtn").addEventListener("click", async () => {
      const body = await api("/api/export", { method: "POST" });
      window.location.href = body.url;
    });

    load().catch(err => setStatus(err.message));
  </script>
</body>
</html>"""


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt: str, *args: Any) -> None:
        print(f"{self.address_string()} - {fmt % args}")

    def send_json(self, payload: Any, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_error_json(self, exc: Exception, status: int = 400) -> None:
        self.send_json({"error": str(exc)}, status)

    def read_json(self) -> dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length)
        return json.loads(raw.decode("utf-8") or "{}")

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/":
                body = INDEX_HTML.encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
            elif parsed.path == "/api/data":
                self.send_json(get_dashboard())
            elif parsed.path.startswith("/exports/"):
                target = (APP_DIR / parsed.path.lstrip("/")).resolve()
                if not str(target).startswith(str(EXPORT_DIR.resolve())) or not target.exists():
                    raise FileNotFoundError("Archivo no encontrado")
                data = target.read_bytes()
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{target.name}"')
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
            else:
                self.send_json({"error": "No encontrado"}, 404)
        except Exception as exc:
            self.send_error_json(exc, 500)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/expense":
                self.send_json(upsert_expense(self.read_json()))
            elif parsed.path == "/api/import":
                form = cgi.FieldStorage(fp=self.rfile, headers=self.headers, environ={"REQUEST_METHOD": "POST"})
                item = form["file"] if "file" in form else None
                if item is None or not getattr(item, "filename", ""):
                    raise ValueError("Elegí un archivo Excel para cargar")
                self.send_json(import_workbook(item.file.read(), item.filename))
            elif parsed.path == "/api/export":
                out = export_xlsx()
                self.send_json({"url": f"/exports/{out.name}"})
            else:
                self.send_json({"error": "No encontrado"}, 404)
        except Exception as exc:
            self.send_error_json(exc, 400)

    def do_DELETE(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/expense":
                params = parse_qs(parsed.query)
                self.send_json(delete_expense(int(params.get("id", ["0"])[0])))
            else:
                self.send_json({"error": "No encontrado"}, 404)
        except Exception as exc:
            self.send_error_json(exc, 400)


def main() -> None:
    init_db()
    seed_default_if_empty()
    host = os.environ.get("MIRAMAR_HOST", "127.0.0.1")
    port = int(os.environ.get("MIRAMAR_PORT", "8765"))
    httpd = ThreadingHTTPServer((host, port), Handler)
    print(f"Miramar app lista en http://{host}:{port}")
    print(f"Base SQLite: {DB_PATH}")
    httpd.serve_forever()


if __name__ == "__main__":
    main()
